import datetime
import io
import os
import json
import pandas as pd
from azure.storage.blob import BlobServiceClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows

# --- Excel Style Definitions ---
# Color styles used throughout the Excel workbook for headers and metric highlighting
HEADER_FILL = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
FILL = PatternFill(start_color="a4f114", end_color="a4f114", fill_type="solid") # Light red for CPU metrics
GROUP_HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # Blue for Group headers


def generate_excel(container_name):
    """
    Main function responsible for:
    1. Connecting to Azure Blob Storage and setting up the container.
    2. Downloading host group information (optional JSON file).
    3. Downloading and processing all CSV metric files into a consolidated structure.
    4. Creating an Excel workbook with Dashboard and All Hosts sheets.
    5. Styling and formatting the data in the Excel sheets.
    6. Uploading the final Excel report and cleaning up CSV files.
    """

    # --- Azure Connection and Setup ---
    connect_str = os.getenv("AZURE_STORAGE_CONNECTION_STRING")
    if not connect_str:
        raise ValueError("AZURE_STORAGE_CONNECTION_STRING is not configured")

    blob_service_client = BlobServiceClient.from_connection_string(connect_str)
    container_client = blob_service_client.get_container_client(container_name)

    # Attempt to create container (it will pass if container already exists)
    try:
        container_client.create_container()
    except:
        pass

    # --- Load Host Group Information (Optional) ---
    try:
        groups_blob = container_client.get_blob_client("_hostgroups_info.json")
        groups_info = json.loads(groups_blob.download_blob().content_as_text())
        host_to_groups = groups_info.get('host_to_groups', {})
        groups_data = groups_info.get('groups', {})
    except:
        print(f"[{container_name}] No host groups info found, continuing without group data")
        host_to_groups = {}
        groups_data = {}

    # --- Download and Process CSV Metric Files ---
    host_count = 0
    csv_data = {}
    csv_blobs_processed = []

    for blob in container_client.list_blobs():
        # ONLY process files ending with .csv and not starting with '_'
        # This prevents re-analyzing old reports or metadata
        if blob.name.lower().endswith(".csv") and not blob.name.startswith("_"):
            blob_client = container_client.get_blob_client(blob.name)
            try:
                stream = io.StringIO(blob_client.download_blob().content_as_text())
                csv_data[blob.name] = pd.read_csv(stream)
                csv_blobs_processed.append(blob.name)
                host_count += 1
            except Exception as e:
                print(f"[{container_name}] Error reading CSV {blob.name}: {e}")

    if not csv_data:
        print(f"[{container_name}] No CSV data found. Skipping Excel generation.")
        return

    # --- Create Workbook and "All Hosts" Sheet ---
    wb = Workbook()
    ws_all = wb.active 
    ws_all.title = "All Hosts"

    headers = ["Host", "Metric", "Min", "Max", "Avg", "Unit", "Samples", "Groups"]
    ws_all.append(headers)

    for col in range(1, len(headers) + 1):
        cell = ws_all.cell(1, col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    row_count = 2
    group_metrics = {} 

    # Populate "All Hosts" sheet and aggregate group metrics
    for csv_name, df in csv_data.items():
        host_name = csv_name.replace(".csv", "")
        groups_str = ";".join(host_to_groups.get(host_name, ["Unknown"]))

        for _, row in df.iterrows():
            unit = row.get('Unit', '')
            ws_all.append([host_name, row['Metric'], row['Min'], row['Max'], row['Avg'], unit, row['Samples'], groups_str])

            for group in host_to_groups.get(host_name, ["Unknown"]):
                group_metrics.setdefault(group, {})
                group_metrics[group].setdefault(host_name, [])
                group_metrics[group][host_name].append({
                    'metric': row['Metric'], 'min': row['Min'], 'max': row['Max'],
                    'avg': row['Avg'], 'unit': unit, 'samples': row['Samples']
                })
            row_count += 1

    # --- Create "Dashboard" Sheet ---
    ws_dashboard = wb.create_sheet("Dashboard", 0)
    ws_dashboard['B2'] = "ZABBIX MONITORING REPORT"
    ws_dashboard['B2'].font = Font(size=16, bold=True, color="2E75B6")
    ws_dashboard.merge_cells('B2:F2')
    ws_dashboard['B3'] = f"Generated: {datetime.datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws_dashboard['B3'].font = Font(size=10, italic=True)

    # Statistics
    stats = [["Total Hosts", host_count], ["Total Metrics", row_count - 2], ["Total Host Groups", len(group_metrics)], ["Period", "Last 30 days"]]
    row_start = 5
    for i, row in enumerate(stats, start=row_start):
        ws_dashboard.cell(i, 2, row[0]).font = Font(bold=True)
        ws_dashboard.cell(i, 4, row[1]).font = Font(size=11)

    # Calculate Global Averages
    global_cpu_values = []
    global_mem_values = []
    for df in csv_data.values():
        for _, row in df.iterrows():
            m = row['Metric'].lower()
            try:
                val = float(row['Avg'])
                if 'cpu' in m and ('util' in m or 'usage' in m): global_cpu_values.append(val)
                elif 'mem' in m and ('utilization' in m or 'pavailable' in m): global_mem_values.append(val)
            except: pass

    ws_dashboard.cell(row_start + len(stats) + 1, 2, "Global CPU Avg (%)").font = Font(bold=True)
    ws_dashboard.cell(row_start + len(stats) + 1, 4, sum(global_cpu_values)/len(global_cpu_values) if global_cpu_values else 0).number_format = '0.00'
    ws_dashboard.cell(row_start + len(stats) + 2, 2, "Global Memory Avg (%)").font = Font(bold=True)
    ws_dashboard.cell(row_start + len(stats) + 2, 4, sum(global_mem_values)/len(global_mem_values) if global_mem_values else 0).number_format = '0.00'

    # Detailed Group Sections (Simplified display)
    group_col_start = 9 
    group_row = 2
    for group_name, hosts_data in sorted(group_metrics.items()):
        ws_dashboard.cell(group_row, group_col_start, f"{group_name}")
        ws_dashboard.cell(group_row, group_col_start).fill = GROUP_HEADER_FILL
        ws_dashboard.cell(group_row, group_col_start).font = Font(color="FFFFFF", bold=True)
        ws_dashboard.merge_cells(start_row=group_row, start_column=group_col_start, end_row=group_row, end_column=group_col_start + 6)
        group_row += 1
        
        headers = ["Host", "Metric", "Min", "Max", "Avg", "Unit", "Samples"]
        for idx, h in enumerate(headers, start=group_col_start):
            cell = ws_dashboard.cell(group_row, idx, h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
        group_row += 1

        for host_name in sorted(hosts_data.keys()):
            for m in hosts_data[host_name]:
                ws_dashboard.cell(group_row, group_col_start, host_name)
                ws_dashboard.cell(group_row, group_col_start + 1, m['metric'])
                ws_dashboard.cell(group_row, group_col_start + 2, m['min'])
                ws_dashboard.cell(group_row, group_col_start + 3, m['max'])
                ws_dashboard.cell(group_row, group_col_start + 4, m['avg'])
                ws_dashboard.cell(group_row, group_col_start + 5, m['unit'])
                ws_dashboard.cell(group_row, group_col_start + 6, m['samples'])
                group_row += 1
        group_row += 2

    # --- Save and Upload Excel File ---
    excel_output = io.BytesIO()
    wb.save(excel_output)
    filename = f"Zabbix_Report_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    container_client.get_blob_client(filename).upload_blob(excel_output.getvalue(), overwrite=True)
    print(f"[{container_name}] Excel '{filename}' uploaded successfully.")

    # --- Cleanup: Delete processed CSV files ---
    print(f"[{container_name}] Cleaning up {len(csv_blobs_processed)} processed CSV files...")
    for b in csv_blobs_processed:
        try:
            container_client.delete_blob(b)
        except Exception as e:
            print(f"[{container_name}] Failed to delete {b}: {e}")


if __name__ == "__main__":
    cname = os.getenv("CONTAINER_NAME", "metrics")
    generate_excel(cname)